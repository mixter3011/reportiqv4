import re
import threading
import pandas as pd
import tkinter as tk
import openpyxl.styles
from pathlib import Path
from openpyxl import Workbook
from tkinter.font import Font
from openpyxl import Workbook
from typing import List, Tuple
from tkinter import ttk, messagebox
from openpyxl.styles import Font, Border, Side
from utils.ops import brw_files, val_file, sim_up
from utils.report import create_portfolio_reports
from utils.processing import ld_data, chk_files, conv
from openpyxl.utils.dataframe import dataframe_to_rows

class ReportIQ:
    def __init__(self, root):
        self.root = root
        self.root.title("ReportIQ")
        self.root.geometry("600x700")  
        self.root.configure(bg='#f8f9fa')  
        
        self.accepted_types = {
            'Excel files': ('*.xlsx', '*.xls', '*.xlsm'),
            'CSV files': ('*.csv',)
        }
        self.required_files = [
            'Portfolio Value.csv', 'Holding.csv', 'XIRR.csv', 
            'Equity.csv', 'Debt.csv', 'FNO.csv', 'Profits.csv'
        ]
        self.files_to_upload: List[Tuple[str, int, str]] = []
        self.desktop_path = Path.home() / "Desktop"
        self.output_dir = self.desktop_path / "converted_files"
        self.output_dir.mkdir(exist_ok=True)
        self.portfolio_dir = self.desktop_path / "portfolio_reports"
        self.portfolio_dir.mkdir(exist_ok=True)
        self.excel_dir = self.desktop_path / "excel_reports"
        self.excel_dir.mkdir(exist_ok=True)
        
        self.style = ttk.Style()
        self.style.configure('TLabel', background='#f8f9fa', font=('Arial', 10))
        self.style.configure('TButton', font=('Arial', 10, 'bold'))
        self.style.configure('TFrame', background='#f8f9fa')
        self.style.configure('DottedFrame.TFrame', borderwidth=2, relief='groove')
        
        self.setup_ui()

    def setup_ui(self):
        self.main_container = ttk.Frame(self.root, padding="20", style='TFrame')
        self.main_container.pack(fill=tk.BOTH, expand=True)
        
        self.create_login_section()
        
        self.create_upload_interface()
        
    def create_login_section(self):
        login_frame = ttk.Frame(self.main_container, style='TFrame')
        login_frame.pack(fill=tk.X, pady=(0, 20))
        
        url_label = ttk.Label(login_frame, text="🔗 Enter URL:", style='TLabel')
        url_label.pack(anchor='w')
        self.url_input = ttk.Entry(login_frame, width=50)
        self.url_input.pack(fill=tk.X, pady=(0, 5))
        
        username_label = ttk.Label(login_frame, text="👤 Enter Username:", style='TLabel')
        username_label.pack(anchor='w')
        self.username_input = ttk.Entry(login_frame, width=50)
        self.username_input.pack(fill=tk.X, pady=(0, 5))
        
        password_label = ttk.Label(login_frame, text="🔒 Enter Password:", style='TLabel')
        password_label.pack(anchor='w')
        self.password_input = ttk.Entry(login_frame, width=50, show="*")
        self.password_input.pack(fill=tk.X, pady=(0, 5))
        
        login_button = ttk.Button(
            login_frame, 
            text="🚀 LOGIN", 
            command=self.login
        )
        login_button.pack(pady=(5, 10), fill=tk.X)
        
        separator = ttk.Separator(login_frame, orient='horizontal')
        separator.pack(fill=tk.X, pady=(10, 10))

    def create_upload_interface(self):
        self.drop_frame = ttk.Frame(
            self.main_container, 
            style='DottedFrame.TFrame'
        )
        self.drop_frame.pack(fill=tk.X, pady=(0, 20))
        
        upload_frame = ttk.Frame(self.drop_frame, style='TFrame')
        upload_frame.pack(pady=20)
        
        upload_icon = ttk.Label(upload_frame, text="↑", style='TLabel')
        upload_icon.pack()
        
        upload_label = ttk.Label(
            upload_frame,
            text="Drag and drop files here or upload",
            font=('Arial', 11),
            style='TLabel'
        )
        upload_label.pack()

        file_types = ttk.Label(
            self.drop_frame,
            text="Accepted file types: Excel (.xlsx, .xls, .xlsm), CSV (.csv)",
            foreground="gray",
            style='TLabel'
        )
        file_types.pack(pady=(0, 10))

        upload_btn = ttk.Button(
            self.drop_frame,
            text="Upload",
            command=self.browse_files
        )
        upload_btn.pack(pady=(0, 20))

        self.files_counter = ttk.Label(
            self.main_container,
            text="0 of 7 files uploaded",
            style='TLabel'
        )
        self.files_counter.pack(anchor=tk.W, pady=(0, 10))

        self.files_frame = ttk.Frame(self.main_container)
        self.files_frame.pack(fill=tk.BOTH, expand=True)

        button_frame = ttk.Frame(self.main_container)
        button_frame.pack(pady=(10, 0))

        self.convert_btn = ttk.Button(
            button_frame,
            text="Convert Files",
            command=self.process_files,
            state="disabled"
        )
        self.convert_btn.pack(side=tk.LEFT, padx=5)

        self.generate_btn = ttk.Button(
            button_frame,
            text="Generate",
            command=self.generate_files,
            state="disabled"
        )
        self.generate_btn.pack(side=tk.LEFT, padx=5)

        self.excel_btn = ttk.Button(
            button_frame,
            text="Generate Excel",
            command=self.generate_excel,
            state="disabled"
        )
        self.excel_btn.pack(side=tk.LEFT, padx=5)

    def login(self):
        url = self.url_input.get().strip()
        username = self.username_input.get().strip()
        password = self.password_input.get().strip()

        if not url or not username or not password:
            messagebox.showwarning("Error", "Please fill in all login fields!")
            return

        # TODO: 
        messagebox.showinfo("Login", "Login functionality to be implemented")

    def generate_excel(self):
        def generate_report():
            try:
                converted_dir = self.output_dir
        
                required_files = {
                    'MF Transaction': converted_dir / 'MF Transaction.csv',
                    'Portfolio Value': converted_dir / 'Portfolio Value.csv',
                    'Equity': converted_dir / 'Equity.csv',
                    'Debt': converted_dir / 'Debt.csv',
                    'Bond': converted_dir / 'Bond.csv'
                }

                missing = [f.name for f in required_files.values() if not f.exists()]
                if missing:
                    raise FileNotFoundError(
                        f"Missing required files: {', '.join(missing)}"
                    )

                mf_df = pd.read_csv(required_files['MF Transaction'], header=None)
                portfolio_df = pd.read_csv(required_files['Portfolio Value'])
                equity_df = pd.read_csv(required_files['Equity'])
                debt_df = pd.read_csv(required_files['Debt'])
                bond_df = pd.read_csv(required_files['Bond'])

                wb = Workbook()
                ws = wb.active
                ws.title = "Portfolio Report"
        
                # Define styles
                bold_font = Font(bold=True)
                header_font = Font(bold=True, size=12)
                title_font = Font(bold=True, size=14)
                thin_border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )

                # Extract client information
                client_line = mf_df.iloc[2, 1]
                parts = [part.strip() for part in client_line.split('/')]
                client_name = parts[-1]
                client_id = parts[0] if len(parts) > 1 else "UNKNOWN"
            
            # Sanitize filename
                safe_name = re.sub(r'[\\/*?:"<>|]', '_', client_name)
                safe_name = safe_name.replace(' ', '_')
                if not safe_name:  # Fallback if empty
                    safe_name = "Unknown_Client"
                
                filename = f"{safe_name}_report.xlsx"
                output_path = self.excel_dir / filename

                # Create header section with light blue background
                for row in range(1, 3):
                    for col in range(1, 15):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = openpyxl.styles.PatternFill(start_color="B6D7E4", end_color="B6D7E4", fill_type="solid")
            
                # Set client name and ID in header
                ws.merge_cells('A1:N1')
                ws.merge_cells('A2:N2')
                ws['A1'] = client_name
                ws['A1'].font = title_font
                ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                ws['A2'] = client_id
                ws['A2'].font = bold_font
                ws['A2'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
                # Get portfolio total value
                portfolio_total = portfolio_df['Portfolio Value'].sum() if 'Portfolio Value' in portfolio_df.columns else 0
            
                # Portfolio Value Section
                row = 4
                ws['A4'] = "Portfolio Value:"
                ws['A4'].font = bold_font
                ws['C4'] = f"{portfolio_total:,.0f}"
            
                # Holdings Section
                row = 5
                ws['A5'] = "Holdings:"
                ws['A5'].font = bold_font
                ws['C5'] = f"{portfolio_total:,.0f}"
            
                # Cash Section
                row = 6
                ws['A6'] = "Available Cash:"
                ws['A6'].font = bold_font
                ws['C6'] = "0"
            
                # Start Equity Section
                row = 7
            
                # Equity Header
                ws['A7'] = "Equity"
                ws['B7'] = "Qty"
                ws['C7'] = "Buy_Price"
                ws['D7'] = "Market_Price"
                ws['E7'] = "P&L"
                ws['F7'] = "Market_Value"
            
            # Debt Header (positioned to the right)
                ws['G7'] = "Debt"
                ws['H7'] = "Qty"
                ws['I7'] = "Buy_Price"
                ws['J7'] = "Market_Price"
                ws['K7'] = "P&L"
                ws['L7'] = "Market_Value"
            
            # Format header row
                for col in range(1, 13):
                    cell = ws.cell(row=7, column=col)
                    cell.font = bold_font
                    cell.border = thin_border
            
            # Calculate equity categories
                direct_equity = equity_df[equity_df['Category'] == 'Direct Equity'] if 'Category' in equity_df.columns else pd.DataFrame()
                equity_etf = equity_df[equity_df['Category'] == 'Equity ETF'] if 'Category' in equity_df.columns else pd.DataFrame()
                equity_mf = equity_df[equity_df['Category'] == 'Equity Mutual Fund'] if 'Category' in equity_df.columns else pd.DataFrame()
            
            # Calculate debt categories
                debt_etf = debt_df[debt_df['Category'] == 'Debt ETF'] if 'Category' in debt_df.columns else pd.DataFrame()
                debt_mf = debt_df[debt_df['Category'] == 'Debt Mutual Fund'] if 'Category' in debt_df.columns else pd.DataFrame()
            
            # Process Total Equity row
                row = 8
                ws['A8'] = "Total Equity"
                ws['A8'].font = bold_font
            
            # Calculate total P&L and Market Value for equity
                total_equity_pl = (direct_equity['P&L'].sum() if 'P&L' in direct_equity.columns else 0) + \
                                (equity_etf['P&L'].sum() if 'P&L' in equity_etf.columns else 0) + \
                                (equity_mf['P&L'].sum() if 'P&L' in equity_mf.columns else 0)
            
                total_equity_mv = (direct_equity['Market Value'].sum() if 'Market Value' in direct_equity.columns else 0) + \
                                (equity_etf['Market Value'].sum() if 'Market Value' in equity_etf.columns else 0) + \
                                (equity_mf['Market Value'].sum() if 'Market Value' in equity_mf.columns else 0)
            
                ws['E8'] = f"{total_equity_pl:,.0f}"
                ws['F8'] = f"{total_equity_mv:,.0f}"
            
                # Calculate equity percentage
                equity_percentage = (total_equity_mv / portfolio_total) * 100 if portfolio_total > 0 else 0
                ws['G8'] = f"{equity_percentage:.2f}%"
            
                # Add cell borders
                for col in range(1, 8):
                    cell = ws.cell(row=8, column=col)
                    cell.border = thin_border
            
            # Process Direct Equity
                row = 9
                ws['A9'] = "Direct Equity"
                ws['A9'].font = bold_font
            
                direct_equity_pl = direct_equity['P&L'].sum() if 'P&L' in direct_equity.columns else 0
                direct_equity_mv = direct_equity['Market Value'].sum() if 'Market Value' in direct_equity.columns else 0
            
                ws['E9'] = f"{direct_equity_pl:,.0f}"
                ws['F9'] = f"{direct_equity_mv:,.0f}"
            
                direct_equity_percentage = (direct_equity_mv / portfolio_total) * 100 if portfolio_total > 0 else 0
                ws['G9'] = f"{direct_equity_percentage:.2f}%"
            
                for col in range(1, 8):
                    cell = ws.cell(row=9, column=col)
                    cell.border = thin_border
            
            # Add Direct Equity holdings
                row = 10
                for idx, stock in direct_equity.iterrows():
                    ws[f'A{row}'] = stock.get('Name', 'Unknown')
                    ws[f'B{row}'] = stock.get('Quantity', 0)
                    ws[f'C{row}'] = stock.get('Buy Price', 0)
                    ws[f'D{row}'] = stock.get('Market Price', 0)
                    ws[f'E{row}'] = stock.get('P&L', 0)
                    ws[f'F{row}'] = stock.get('Market Value', 0)
                
                # Add borders
                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                
                    row += 1
            
            # Process Equity ETF
                ws[f'A{row}'] = "Equity ETF"
                ws[f'A{row}'].font = bold_font
            
                etf_pl = equity_etf['P&L'].sum() if 'P&L' in equity_etf.columns else 0
                etf_mv = equity_etf['Market Value'].sum() if 'Market Value' in equity_etf.columns else 0
            
                ws[f'E{row}'] = f"{etf_pl:,.0f}"
                ws[f'F{row}'] = f"{etf_mv:,.0f}"
            
                etf_percentage = (etf_mv / portfolio_total) * 100 if portfolio_total > 0 else 0
                ws[f'G{row}'] = f"{etf_percentage:.2f}%"
            
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
            
                row += 1
            
            # Add ETF holdings
                for idx, etf in equity_etf.iterrows():
                    ws[f'A{row}'] = etf.get('Name', 'Unknown')
                    ws[f'B{row}'] = etf.get('Quantity', 0)
                    ws[f'C{row}'] = etf.get('Buy Price', 0)
                    ws[f'D{row}'] = etf.get('Market Price', 0)
                    ws[f'E{row}'] = etf.get('P&L', 0)
                    ws[f'F{row}'] = etf.get('Market Value', 0)
                
                # Add borders
                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                
                    row += 1
            
            # Process Equity Mutual Fund
                ws[f'A{row}'] = "Equity Mutual Fund"
                ws[f'A{row}'].font = bold_font
            
                mf_pl = equity_mf['P&L'].sum() if 'P&L' in equity_mf.columns else 0
                mf_mv = equity_mf['Market Value'].sum() if 'Market Value' in equity_mf.columns else 0
            
                ws[f'E{row}'] = f"{mf_pl:,.0f}"
                ws[f'F{row}'] = f"{mf_mv:,.0f}"
            
                mf_percentage = (mf_mv / portfolio_total) * 100 if portfolio_total > 0 else 0
                ws[f'G{row}'] = f"{mf_percentage:.2f}%"
            
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
            
                row += 1
            
            # Add Mutual Fund holdings
                for idx, mf in equity_mf.iterrows():
                    ws[f'A{row}'] = mf.get('Name', 'Unknown')
                    ws[f'B{row}'] = mf.get('Quantity', 0)
                    ws[f'C{row}'] = mf.get('Buy Price', 0)
                    ws[f'D{row}'] = mf.get('Market Price', 0)
                    ws[f'E{row}'] = mf.get('P&L', 0)
                    ws[f'F{row}'] = mf.get('Market Value', 0)
                
                # Add borders
                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                
                    row += 1
            
            # Add Debt section
            # Reset to row 8 for debt section (parallel to equity)
                debt_row = 8
            
            # Process Cash Equivalent (Debt)
                ws['G8'] = "Cash Equivalent"
                ws['G8'].font = bold_font
            
            # Calculate total debt values
                total_debt_mv = (debt_etf['Market Value'].sum() if 'Market Value' in debt_etf.columns else 0) + \
                            (debt_mf['Market Value'].sum() if 'Market Value' in debt_mf.columns else 0)
            
                total_debt_pl = (debt_etf['P&L'].sum() if 'P&L' in debt_etf.columns else 0) + \
                            (debt_mf['P&L'].sum() if 'P&L' in debt_mf.columns else 0)
            
                debt_percentage = (total_debt_mv / portfolio_total) * 100 if portfolio_total > 0 else 0
            
                ws['K8'] = f"{total_debt_pl:,.0f}"
                ws['L8'] = f"{total_debt_mv:,.0f}"
                ws['M8'] = f"{debt_percentage:.2f}%"
            
                for col in range(7, 14):
                    cell = ws.cell(row=8, column=col)
                    cell.border = thin_border
            
            # Process Debt ETF
                debt_row = 9
                ws['G9'] = "Debt ETF"
                ws['G9'].font = bold_font
            
                debt_etf_pl = debt_etf['P&L'].sum() if 'P&L' in debt_etf.columns else 0
                debt_etf_mv = debt_etf['Market Value'].sum() if 'Market Value' in debt_etf.columns else 0
            
                debt_etf_percentage = (debt_etf_mv / portfolio_total) * 100 if portfolio_total > 0 else 0
            
                ws['K9'] = f"{debt_etf_pl:,.0f}"
                ws['L9'] = f"{debt_etf_mv:,.0f}"
                ws['M9'] = f"{debt_etf_percentage:.2f}%"
            
                for col in range(7, 14):
                    cell = ws.cell(row=9, column=col)
                    cell.border = thin_border
            
            # Add Debt ETF holdings
                debt_row = 10
                for idx, etf in debt_etf.iterrows():
                    ws[f'G{debt_row}'] = etf.get('Name', 'Unknown')
                    ws[f'H{debt_row}'] = etf.get('Quantity', 0)
                    ws[f'I{debt_row}'] = etf.get('Buy Price', 0)
                    ws[f'J{debt_row}'] = etf.get('Market Price', 0)
                    ws[f'K{debt_row}'] = etf.get('P&L', 0)
                    ws[f'L{debt_row}'] = etf.get('Market Value', 0)
                
                # Add borders
                    for col in range(7, 13):
                        cell = ws.cell(row=debt_row, column=col)
                        cell.border = thin_border
                
                    debt_row += 1
            
            # Process Debt Mutual Fund
                ws[f'G{debt_row}'] = "Debt Mutual Fund"
                ws[f'G{debt_row}'].font = bold_font
            
                debt_mf_pl = debt_mf['P&L'].sum() if 'P&L' in debt_mf.columns else 0
                debt_mf_mv = debt_mf['Market Value'].sum() if 'Market Value' in debt_mf.columns else 0
            
                debt_mf_percentage = (debt_mf_mv / portfolio_total) * 100 if portfolio_total > 0 else 0
            
                ws[f'K{debt_row}'] = f"{debt_mf_pl:,.0f}"
                ws[f'L{debt_row}'] = f"{debt_mf_mv:,.0f}"
                ws[f'M{debt_row}'] = f"{debt_mf_percentage:.2f}%"
            
                for col in range(7, 14):
                    cell = ws.cell(row=debt_row, column=col)
                    cell.border = thin_border
            
                debt_row += 1
            
            # Add Debt Mutual Fund holdings
                for idx, mf in debt_mf.iterrows():
                    ws[f'G{debt_row}'] = mf.get('Name', 'Unknown')
                    ws[f'H{debt_row}'] = mf.get('Quantity', 0)
                    ws[f'I{debt_row}'] = mf.get('Buy Price', 0)
                    ws[f'J{debt_row}'] = mf.get('Market Price', 0)
                    ws[f'K{debt_row}'] = mf.get('P&L', 0)
                    ws[f'L{debt_row}'] = mf.get('Market Value', 0)
                
                # Add borders
                    for col in range(7, 13):
                        cell = ws.cell(row=debt_row, column=col)
                        cell.border = thin_border
                
                    debt_row += 1
            
            # Calculate the row for Grand Total in debt section
                ws[f'G{debt_row}'] = "Grand Total"
                ws[f'G{debt_row}'].font = bold_font
                ws[f'L{debt_row}'] = f"{total_debt_mv:,.0f}"
            
            # Add Bond section
            # Find the maximum row between equity and debt sections
                bond_start_row = max(row, debt_row) + 2
            
            # Add Bond header with light blue background
                for col in range(1, 7):
                    cell = ws.cell(row=bond_start_row, column=col)
                    cell.fill = openpyxl.styles.PatternFill(start_color="B6D7E4", end_color="B6D7E4", fill_type="solid")
            
                ws[f'A{bond_start_row}'] = "Bond"
                ws[f'A{bond_start_row}'].font = bold_font
            
            # Bond table headers
                bond_start_row += 1
                ws[f'A{bond_start_row}'] = "Bond"
                ws[f'B{bond_start_row}'] = "Qty"
                ws[f'C{bond_start_row}'] = "Buy_Price"
                ws[f'D{bond_start_row}'] = "Market_Price"
                ws[f'E{bond_start_row}'] = "P&L"
                ws[f'F{bond_start_row}'] = "Market_Value"
            
            # Format Bond headers
                for col in range(1, 7):
                    cell = ws.cell(row=bond_start_row, column=col)
                    cell.font = bold_font
                    cell.border = thin_border
            
                bond_start_row += 1
            
            # Process Cash Equivalent (Bond)
                ws[f'A{bond_start_row}'] = "Cash Equivalent"
                ws[f'A{bond_start_row}'].font = bold_font
            
                total_bond_mv = bond_df['Market Value'].sum() if 'Market Value' in bond_df.columns else 0
                total_bond_pl = bond_df['P&L'].sum() if 'P&L' in bond_df.columns else 0
            
                bond_percentage = (total_bond_mv / portfolio_total) * 100 if portfolio_total > 0 else 0
            
                ws[f'E{bond_start_row}'] = f"{total_bond_pl:,.0f}"
                ws[f'F{bond_start_row}'] = f"{total_bond_mv:,.0f}"
                ws[f'G{bond_start_row}'] = f"{bond_percentage:.2f}%"
            
                for col in range(1, 8):
                    cell = ws.cell(row=bond_start_row, column=col)
                    cell.border = thin_border
            
                bond_start_row += 1
            
            # Add Bond holdings
                for idx, bond in bond_df.iterrows():
                    ws[f'A{bond_start_row}'] = bond.get('Name', 'Unknown')
                    ws[f'B{bond_start_row}'] = bond.get('Quantity', 0)
                    ws[f'C{bond_start_row}'] = bond.get('Buy Price', 0)
                    ws[f'D{bond_start_row}'] = bond.get('Market Price', 0)
                    ws[f'E{bond_start_row}'] = bond.get('P&L', 0)
                    ws[f'F{bond_start_row}'] = bond.get('Market Value', 0)
                
                # Add borders
                    for col in range(1, 7):
                        cell = ws.cell(row=bond_start_row, column=col)
                        cell.border = thin_border
                
                    bond_start_row += 1
            
            # Grand Total for Bond
                ws[f'A{bond_start_row}'] = "Grand Total"
                ws[f'A{bond_start_row}'].font = bold_font
                ws[f'F{bond_start_row}'] = f"{total_bond_mv:,.0f}"
            
                for col in range(1, 7):
                    cell = ws.cell(row=bond_start_row, column=col)
                    cell.border = thin_border
            
            # Adjust column widths
                for col in range(1, 14):
                    column_letter = openpyxl.utils.get_column_letter(col)
                    ws.column_dimensions[column_letter].width = 15
            
            # First column needs to be wider for names
                ws.column_dimensions['A'].width = 25
            
            # Save the workbook
                wb.save(output_path)
                self.root.after(0, lambda: messagebox.showinfo(
                    "Success", 
                    f"Excel report generated successfully at:\n{output_path}"
                ))

            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror(
                    "Error",
                    f"Failed to generate Excel report: {str(e)}"
                ))

    # Make sure to add the import for openpyxl.styles.PatternFill at the top of the file
    # import openpyxl.styles

        threading.Thread(target=generate_report, daemon=True).start()

    def browse_files(self):
        files = brw_files(self.root, self.accepted_types)
        if files:
            self.add_files(files)

    def add_files(self, files):
        for file in files:
            if len(self.files_to_upload) >= 7:
                messagebox.showwarning("Limit Reached", "Maximum 7 files can be uploaded at once.")
                break
                
            if not val_file(file):
                messagebox.showerror("Invalid File", f"File type not supported: {file}")
                continue
                
            filename = Path(file).name
            size = Path(file).stat().st_size // (1024 * 1024)
            
            file_frame = ttk.Frame(self.files_frame)
            file_frame.pack(fill=tk.X, pady=5)
            
            icon_text = "📊" if filename.endswith('.csv') else "📑"
            icon_label = ttk.Label(file_frame, text=icon_text, style='TLabel')
            icon_label.pack(side=tk.LEFT, padx=5)
            
            name_label = ttk.Label(
                file_frame,
                text=f"{filename} {size} mb",
                style='TLabel'
            )
            name_label.pack(side=tk.LEFT, padx=5)
            
            progress = ttk.Progressbar(
                file_frame,
                length=200,
                mode='determinate'
            )
            progress.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
            
            cancel_btn = ttk.Button(
                file_frame,
                text="✕",
                width=3,
                command=lambda f=file_frame: self.remove_file(f)
            )
            cancel_btn.pack(side=tk.RIGHT, padx=5)
            
            self.files_to_upload.append((filename, size, file))
            self.update_counter()
            
            sim_up(progress, self.root)
        
        button_state = 'normal' if self.files_to_upload else 'disabled'
        self.convert_btn['state'] = button_state
        self.generate_btn['state'] = button_state
        self.excel_btn['state'] = button_state

    def remove_file(self, file_frame):
        index = list(self.files_frame.children.values()).index(file_frame)
        self.files_to_upload.pop(index)
        file_frame.destroy()
        self.update_counter()
        button_state = 'normal' if self.files_to_upload else 'disabled'
        self.convert_btn['state'] = button_state
        self.generate_btn['state'] = button_state
        self.excel_btn['state'] = button_state

    def update_counter(self):
        self.files_counter.configure(
            text=f"{len(self.files_to_upload)} of 7 files uploaded"
        )

    def process_files(self):
        self.convert_btn['state'] = 'disabled'
        self.generate_btn['state'] = 'disabled'
        self.excel_btn['state'] = 'disabled'
        
        def conversion_thread():
            converted_files = []
            for filename, _, file_path in self.files_to_upload:
                try:
                    if file_path.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                        converted = conv(file_path, self.output_dir)
                        converted_files.extend(converted)
                    else:
                        output_path = self.output_dir / filename
                        pd.read_csv(file_path).to_csv(output_path, index=False)
                        converted_files.append(filename)
                        
                except Exception as e:
                    messagebox.showerror("Error", f"Error processing {filename}: {str(e)}")
                    continue
            
            messagebox.showinfo("Success", "Files have been processed successfully!")
            self.convert_btn['state'] = 'normal'
            self.generate_btn['state'] = 'normal'
            self.excel_btn['state'] = 'normal'

        thread = threading.Thread(target=conversion_thread)
        thread.start()

    def generate_files(self):
        if not chk_files(self.files_to_upload, self.required_files, self.output_dir):
            return

        self.generate_btn['state'] = 'disabled'
        self.convert_btn['state'] = 'disabled'
        self.excel_btn['state'] = 'disabled'

        progress_window = tk.Toplevel(self.root)
        progress_window.title("Generating Reports")
        progress_window.geometry("300x150")
        progress_window.transient(self.root)
    
        progress_label = ttk.Label(progress_window, text="Generating portfolio reports...")
        progress_label.pack(pady=20)
    
        progress_bar = ttk.Progressbar(progress_window, mode='indeterminate')
        progress_bar.pack(pady=10, padx=20, fill=tk.X)
        progress_bar.start()

        def generation_thread():
            try:
                data = ld_data(self.files_to_upload, self.required_files, self.output_dir)
                if not data:
                    progress_window.destroy()
                    return

                self.root.after(0, lambda: self.create_reports_on_main_thread(data, progress_window))
            
            except Exception as e:
                def error_handler():
                    progress_window.destroy()
                    messagebox.showerror("Error", f"Error generating reports: {str(e)}")
                    self.reset_interface()
            
                self.root.after(0, error_handler)

        thread = threading.Thread(target=generation_thread)
        thread.start()

    def create_reports_on_main_thread(self, data, progress_window):
        try:
            create_portfolio_reports(data, self.portfolio_dir)
            progress_window.destroy()
            messagebox.showinfo(
                "Success", 
                f"Portfolio reports have been generated successfully in the '{self.portfolio_dir}' directory."
            )
            self.reset_interface()
        except Exception as e:
            messagebox.showerror("Error", f"Error generating reports: {str(e)}")
            self.reset_interface()

    def reset_interface(self):
        self.files_to_upload = []
        self.create_upload_interface()
        self.root.update()
