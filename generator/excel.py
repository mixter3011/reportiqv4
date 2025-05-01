import pandas as pd
from openpyxl import Workbook
from utils.format import format_num, isn
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.formula.translate import Translator

def excel_generator(df):
    client_info_row = df[df['Unnamed: 0'] == 'Client Equity Code/UCID/Name'].index[0]
    client_info = str(df.iloc[client_info_row, 1]).strip()
    parts = client_info.split('/')
    client_code = parts[0].strip()  
    client_name = parts[-1].strip()  

    equity_row = df[df['Unnamed: 0'] == 'Equity:-'].index[0]
    mf_row = df[df['Unnamed: 0'] == 'Mutual Fund:-'].index[0]
    fno_row = df[df['Unnamed: 0'] == 'FnO:-'].index[0]
    bond_row = df[df['Unnamed: 0'] == 'Bond:-'].index[0]

    equity_header = df.iloc[equity_row + 1].tolist()
    equity_end = mf_row - 4  
    equity_data = df.iloc[equity_row + 2:equity_end].copy()

    mf_header = df.iloc[mf_row + 1].tolist()
    mf_end = fno_row - 4
    mf_data = df.iloc[mf_row + 2:mf_end].copy()

    bond_header = df.iloc[bond_row + 1].tolist()
    bond_end = len(df)
    for i in range(bond_row + 2, len(df)):
        if i >= len(df) or pd.isna(df.iloc[i, 0]) or df.iloc[i, 0] == '':
            bond_end = i
            break
    bond_data = df.iloc[bond_row + 2:bond_end].copy()

    equity_data = equity_data[equity_data['Unnamed: 0'] != 'Total:']
    mf_data = mf_data[mf_data['Unnamed: 0'] != 'Total:']
    bond_data = bond_data[bond_data['Unnamed: 0'] != 'Total:']

    wb = Workbook()
    ws = wb.active

    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    ws.merge_cells('A1:O1')
    ws.merge_cells('A2:O2')

    market_value_cells = {}

    ws.cell(row=1, column=1, value=client_name)
    ws.cell(row=1, column=1).fill = light_blue_fill
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=client_code)
    ws.cell(row=2, column=1).fill = light_blue_fill
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    direct_equity = equity_data[~equity_data['Unnamed: 0'].str.contains('ETF', na=False)]
    direct_equity = direct_equity[~direct_equity['Unnamed: 0'].str.contains('Nifty 1D Rate Liquid BeES', na=False)]

    equity_cols_to_keep = [0, 1, 2, 4, 10, 5]  
    equity_rename = {2: "Buy Price", 10: "P&L"}

    
    direct_equity_total = ['Total:']
    
    direct_equity_total.extend([''] * 3)

    
    market_value_col = equity_cols_to_keep[4]  
    market_values = pd.to_numeric(direct_equity[f'Unnamed: {market_value_col}'], errors='coerce')
    direct_equity_market_value = market_values.sum() if not market_values.isna().all() else 0
    direct_equity_total.append(direct_equity_market_value)

    
    pnl_col = equity_cols_to_keep[5]  
    pnl_values = pd.to_numeric(direct_equity[f'Unnamed: {pnl_col}'], errors='coerce')
    direct_equity_pnl = pnl_values.sum() if not pnl_values.isna().all() else 0
    direct_equity_total.append(direct_equity_pnl)

    etf_equity = equity_data[equity_data['Unnamed: 0'].str.contains('ETF', na=False)]
    etf_equity = etf_equity[~etf_equity['Unnamed: 0'].str.contains('Nifty 1D Rate Liquid BeES', na=False)]
    etf_equity = etf_equity[~etf_equity['Unnamed: 0'].str.contains('Nippon India ETF Nifty 8-13 yr G-Sec LongTerm Gilt', na=False)]
    etf_equity = etf_equity[~etf_equity['Unnamed: 0'].str.contains('Nippon India ETF Nifty 5 Yr Benchmark GSec', na=False)]

    
    etf_equity_total = ['Total:']
    
    etf_equity_total.extend([''] * 3)

    
    market_values = pd.to_numeric(etf_equity[f'Unnamed: {market_value_col}'], errors='coerce')
    etf_equity_market_value = market_values.sum() if not market_values.isna().all() else 0
    etf_equity_total.append(etf_equity_market_value)

    
    pnl_values = pd.to_numeric(etf_equity[f'Unnamed: {pnl_col}'], errors='coerce')
    etf_equity_pnl = pnl_values.sum() if not pnl_values.isna().all() else 0
    etf_equity_total.append(etf_equity_pnl)

    debt_etf = equity_data[equity_data['Unnamed: 0'].str.contains('Nifty 1D Rate Liquid BeES', na=False)]
    nifty_5yr_etf = equity_data[equity_data['Unnamed: 0'].str.contains('Nippon India ETF Nifty 5 Yr Benchmark GSec', na=False)]

    if not nifty_5yr_etf.empty:
        debt_etf = pd.concat([debt_etf, nifty_5yr_etf], ignore_index=True)
    
    debt_etf_total = ['Total:']
    
    debt_etf_total.extend([''] * 3)

    
    market_values = pd.to_numeric(debt_etf[f'Unnamed: {market_value_col}'], errors='coerce')
    debt_etf_market_value = market_values.sum() if not market_values.isna().all() else 0
    debt_etf_total.append(debt_etf_market_value)

    
    pnl_values = pd.to_numeric(debt_etf[f'Unnamed: {pnl_col}'], errors='coerce')
    debt_etf_pnl = pnl_values.sum() if not pnl_values.isna().all() else 0
    debt_etf_total.append(debt_etf_pnl)

    gilt_etf = equity_data[equity_data['Unnamed: 0'].str.contains('Nippon India ETF Nifty 8-13 yr G-Sec LongTerm Gilt', na=False)]

    mf_cols_to_keep = [1, 2, 3, 5, 12, 6]  
    mf_rename = {3: "Buy Price", 12: "P&L"}

    equity_mf = mf_data[mf_data['Unnamed: 0'] == 'Equity']

    
    equity_mf_total = ['Total:']
    
    equity_mf_total.extend([''] * 3)

    
    mf_market_value_col = mf_cols_to_keep[4]  
    market_values = pd.to_numeric(equity_mf[f'Unnamed: {mf_market_value_col}'], errors='coerce')
    equity_mf_market_value = market_values.sum() if not market_values.isna().all() else 0
    equity_mf_total.append(equity_mf_market_value)

    
    mf_pnl_col = mf_cols_to_keep[5]  
    pnl_values = pd.to_numeric(equity_mf[f'Unnamed: {mf_pnl_col}'], errors='coerce')
    equity_mf_pnl = pnl_values.sum() if not pnl_values.isna().all() else 0
    equity_mf_total.append(equity_mf_pnl)

    debt_mf = mf_data[mf_data['Unnamed: 0'] == 'Debt']

    if not gilt_etf.empty:
        gilt_for_debt = pd.DataFrame(columns=debt_mf.columns)

        column_mapping = {
            'Unnamed: 0': 'Unnamed: 1',  
            'Unnamed: 1': 'Unnamed: 2',  
            'Unnamed: 2': 'Unnamed: 3',  
            'Unnamed: 4': 'Unnamed: 5',  
            'Unnamed: 5': 'Unnamed: 6',  
            'Unnamed: 10': 'Unnamed: 12' 
        }
    
        for old_col, new_col in column_mapping.items():
            if old_col in gilt_etf.columns and new_col in debt_mf.columns:
                gilt_for_debt[new_col] = gilt_etf[old_col].values
    
        debt_mf = pd.concat([debt_mf, gilt_for_debt], ignore_index=True)

    
    debt_mf_total = ['Total:']
    
    debt_mf_total.extend([''] * 3)

    
    market_values = pd.to_numeric(debt_mf[f'Unnamed: {mf_market_value_col}'], errors='coerce')
    debt_mf_market_value = market_values.sum() if not market_values.isna().all() else 0
    debt_mf_total.append(debt_mf_market_value)

    
    pnl_values = pd.to_numeric(debt_mf[f'Unnamed: {mf_pnl_col}'], errors='coerce')
    debt_mf_pnl = pnl_values.sum() if not pnl_values.isna().all() else 0
    debt_mf_total.append(debt_mf_pnl)

    bond_cols_to_keep = [0, 1, 2, 4, 10, 5]  
    bond_rename = {2: "Buy Price", 10: "P&L"}

    
    bond_total = ['Total:']
    
    bond_total.extend([''] * 3)

    
    bond_market_value_col = bond_cols_to_keep[4]  
    market_values = pd.to_numeric(bond_data[f'Unnamed: {bond_market_value_col}'], errors='coerce')
    bond_market_value = market_values.sum() if not market_values.isna().all() else 0
    bond_total.append(bond_market_value)

    
    bond_pnl_col = bond_cols_to_keep[5]  
    pnl_values = pd.to_numeric(bond_data[f'Unnamed: {bond_pnl_col}'], errors='coerce')
    bond_pnl = pnl_values.sum() if not pnl_values.isna().all() else 0
    bond_total.append(bond_pnl)

    market_value_total_rows = {}

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    portfolio_value_fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid") 
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    row = 5
    ws.cell(row=row + 2, column=1, value="EQUITY").font = Font(bold=True)

    row_portfolio = row - 1  
    row_cash = row_portfolio + 1
    
    ws.merge_cells(f'A{row_portfolio}')
    ws.cell(row=row_portfolio, column=1, value="Portfolio Value").font = Font(bold=True)
    ws.cell(row=row_portfolio, column=1).border = thin_border
    ws.cell(row=row_portfolio, column=1).alignment = Alignment(horizontal="left")

    ws.merge_cells(f'B{row_portfolio}')
    ws.cell(row=row_portfolio, column=2, value=0)
    ws.cell(row=row_portfolio, column=2).border = thin_border
    ws.cell(row=row_portfolio, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row_portfolio, column=2).font = Font(bold=True)
    ws.cell(row=row_portfolio, column=2).number_format = '#,##,##0'
    
    ws.merge_cells(f'A{row_cash}')
    ws.cell(row=row_cash, column=1, value="Cash Data").font = Font(bold=True)
    ws.cell(row=row_cash, column=1).border = thin_border
    ws.cell(row=row_cash, column=1).alignment = Alignment(horizontal="left")
    
    ws.merge_cells(f'B{row_cash}')
    ws.cell(row=row_cash, column=2, value=0)  
    ws.cell(row=row_cash, column=2).border = thin_border
    ws.cell(row=row_cash, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row_cash, column=2).number_format = '#,##,##0'
    
    ws.row_dimensions[row_cash].height = 12
    
    row += 3
    ws.cell(row=row, column=1, value="Direct Equity").font = Font(bold=True)
    row += 1

    equity_col_map = {}
    for new_idx, old_idx in enumerate(equity_cols_to_keep, 1):
        equity_col_map[old_idx] = new_idx

    for old_idx in equity_cols_to_keep:
        new_idx = equity_col_map[old_idx]
        header = equity_header[old_idx]
    
        if old_idx in equity_rename:
            header = equity_rename[old_idx]
        
        ws.cell(row=row, column=new_idx, value=header)
        ws.cell(row=row, column=new_idx).fill = header_fill
        ws.cell(row=row, column=new_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=new_idx).border = thin_border
        ws.column_dimensions[get_column_letter(new_idx)].width = 11

    alloc_col = len(equity_cols_to_keep) + 1
    ws.cell(row=row, column=alloc_col, value="% Alloc")
    ws.cell(row=row, column=alloc_col).fill = header_fill
    ws.cell(row=row, column=alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=alloc_col).border = thin_border
    ws.column_dimensions[get_column_letter(alloc_col)].width = 8

    row += 1

    for _, data_row in direct_equity.iterrows():
        for old_idx in equity_cols_to_keep:
            new_idx = equity_col_map[old_idx]
            value = data_row[old_idx]
            if not pd.isna(value):
                cell = ws.cell(row=row, column=new_idx, value=value)
                if new_idx in [5, 6]:
                    cell.number_format = '##,##,##0'
                if old_idx == 0:
                    ws.cell(row=row, column=new_idx).alignment = Alignment(horizontal="left")
                else:
                    ws.cell(row=row, column=new_idx).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=new_idx).border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=1).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  

    for col_idx in [2, 3, 4, 6]:
        ws.column_dimensions[get_column_letter(col_idx)].width = 11.5
        if col_idx == 6:
            ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=col_idx).border = thin_border
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    market_value_col = 6
    market_value_col_letter = get_column_letter(market_value_col)

    if len(direct_equity) > 0:
        first_data_row = row - len(direct_equity)
        last_data_row = row - 1
    
        cell_references = [f"{market_value_col_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        formula = "=" + "+".join(cell_references)
    
        if not cell_references:
            formula = "=0"
    else:
        formula = "=0"

    market_value_cell = ws.cell(row=row, column=market_value_col)
    market_value_cell.value = formula
    market_value_cell.number_format = '#,##,##0'
    market_value_cell.border = thin_border
    market_value_cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    pnl_col = 5
    pnl_col_letter = get_column_letter(pnl_col)

    if len(direct_equity) > 0:
        first_data_row = row - len(direct_equity)
        last_data_row = row - 1
    
        cell_references = [f"{pnl_col_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        formula = "=" + "+".join(cell_references)
    
        if not cell_references:
            formula = "=0"
    else:
        formula = "=0"

    direct_equity_total_row = row
    market_value_total_rows['direct_equity'] = (market_value_col, row)
    
    pnl_cell = ws.cell(row=row, column=pnl_col)
    pnl_cell.value = formula
    pnl_cell.number_format = '#,##,##0'
    pnl_cell.border = thin_border
    pnl_cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    total_portfolio_value_cell = "B4"  
    alloc_formula = f"={market_value_col_letter}{row}/{total_portfolio_value_cell}*100"

    alloc_cell = ws.cell(row=row, column=alloc_col)
    alloc_cell.value = alloc_formula
    alloc_cell.alignment = Alignment(horizontal="center")
    alloc_cell.border = thin_border
    alloc_cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    alloc_cell.number_format = "0.00\%"

    
    row += 1

    row += 1
    ws.cell(row=row, column=1, value="Equity ETF").font = Font(bold=True)
    row += 1

    for old_idx in equity_cols_to_keep:
        new_idx = equity_col_map[old_idx]
        header = equity_header[old_idx]
    
        if old_idx in equity_rename:
            header = equity_rename[old_idx]
        
        ws.cell(row=row, column=new_idx, value=header)
        ws.cell(row=row, column=new_idx).fill = header_fill
        ws.cell(row=row, column=new_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=new_idx).border = thin_border

    ws.cell(row=row, column=alloc_col, value="% Alloc")
    ws.cell(row=row, column=alloc_col).fill = header_fill
    ws.cell(row=row, column=alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=alloc_col).border = thin_border
    
    row += 1

    for _, data_row in etf_equity.iterrows():
        for old_idx in equity_cols_to_keep:
            new_idx = equity_col_map[old_idx]
            value = data_row[old_idx]
            if not pd.isna(value):
                cell_value = format_num(value)
                ws.cell(row=row, column=new_idx, value=cell_value)
                if old_idx == 0:
                    ws.cell(row=row, column=new_idx).alignment = Alignment(horizontal="left")
                else:
                    ws.cell(row=row, column=new_idx).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=new_idx).border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=1).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    for col_idx in [2, 3, 4]:
        ws.cell(row=row, column=col_idx).border = thin_border
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    first_data_row = row - len(etf_equity)
    last_data_row = row - 1

    market_value_col_letter = get_column_letter(5)
    if first_data_row <= last_data_row:
        cell_references = [f"{market_value_col_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        formula = "=" + "+".join(cell_references)
    else:
        formula = "=0"

    ws.cell(row=row, column=5, value=formula)
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=5).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=5).number_format = '#,##,##0'

    pnl_col_letter = get_column_letter(6)
    if first_data_row <= last_data_row:
        cell_references = [f"{pnl_col_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        formula = "=" + "+".join(cell_references)
    else:
        formula = "=0"

    etf_equity_total_row = row
    market_value_total_rows['etf_equity'] = (6, row)
    
    ws.cell(row=row, column=6, value=formula)
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=6).border = thin_border
    ws.cell(row=row, column=6).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=6).number_format = '#,##,##0'

    ws.cell(row=row, column=alloc_col, value=f"={market_value_col_letter}{row}/B4*100")
    ws.cell(row=row, column=alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=alloc_col).border = thin_border
    ws.cell(row=row, column=alloc_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=alloc_col).number_format = "0.00\%"    
        
    row += 1

    row += 1
    ws.cell(row=row, column=1, value="Equity Mutual Fund").font = Font(bold=True)
    row += 1

    mf_col_map = {}
    for new_idx, old_idx in enumerate(mf_cols_to_keep, 1):
        mf_col_map[old_idx] = new_idx

    for old_idx in mf_cols_to_keep:
        new_idx = mf_col_map[old_idx]
        header = mf_header[old_idx]
    
        if old_idx in mf_rename:
            header = mf_rename[old_idx]
        
        ws.cell(row=row, column=new_idx, value=header)
        ws.cell(row=row, column=new_idx).fill = header_fill
        ws.cell(row=row, column=new_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=new_idx).border = thin_border

    ws.cell(row=row, column=alloc_col, value="% Alloc")
    ws.cell(row=row, column=alloc_col).fill = header_fill
    ws.cell(row=row, column=alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=alloc_col).border = thin_border

    row += 1

    for _, data_row in equity_mf.iterrows():
        for old_idx in mf_cols_to_keep:
            new_idx = mf_col_map[old_idx]
            value = data_row[old_idx]
            if not pd.isna(value):
                cell_value = format_num(value)
                ws.cell(row=row, column=new_idx, value=cell_value)
                if old_idx == 1:
                    ws.cell(row=row, column=new_idx).alignment = Alignment(horizontal="left")
                else:
                    ws.cell(row=row, column=new_idx).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=new_idx).border = thin_border
        row += 1
    
    ws.cell(row=row, column=1, value="Total:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=1).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    for col_idx in [2, 3, 4]:
        ws.cell(row=row, column=col_idx).border = thin_border
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    mf_market_value_col_letter = get_column_letter(5)
    mf_pnl_col_letter = get_column_letter(6)

    first_data_row = row - len(equity_mf)
    last_data_row = row - 1

    if first_data_row <= last_data_row:
        market_value_cell_references = [f"{mf_market_value_col_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        market_value_formula = "=" + "+".join(market_value_cell_references)
    else:
        market_value_formula = "=0"

    equity_mf_total_row = row
    market_value_total_rows['equity_mf'] = (6, row)
    
    ws.cell(row=row, column=5, value=market_value_formula)
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=5).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=5).number_format = '#,##,##0'

    if first_data_row <= last_data_row:
        pnl_cell_references = [f"{mf_pnl_col_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        pnl_formula = "=" + "+".join(pnl_cell_references)
    else:
        pnl_formula = "=0"

    ws.cell(row=row, column=6, value=pnl_formula)
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=6).border = thin_border
    ws.cell(row=row, column=6).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=6).number_format = '#,##,##0'

    ws.cell(row=row, column=alloc_col, value=f"={mf_market_value_col_letter}{row}/B4*100")
    ws.cell(row=row, column=alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=alloc_col).border = thin_border
    ws.cell(row=row, column=alloc_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=alloc_col).number_format = "0.00\%"    
    
    row += 1
    
    col_offset = 9  
    row = 7
    ws.cell(row=row, column=col_offset, value="DEBT").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=col_offset, value="Debt ETF").font = Font(bold=True)
    row += 1

    for old_idx in equity_cols_to_keep:
        new_idx = equity_col_map[old_idx]
        header = equity_header[old_idx]
    
        if old_idx in equity_rename:
            header = equity_rename[old_idx]
        
        ws.cell(row=row, column=new_idx + col_offset - 1, value=header)
        ws.cell(row=row, column=new_idx + col_offset - 1).fill = header_fill
        ws.cell(row=row, column=new_idx + col_offset - 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=new_idx + col_offset - 1).border = thin_border
        if old_idx == 0:
            ws.column_dimensions[get_column_letter(new_idx + col_offset - 1)].width = 25
        else:
            ws.column_dimensions[get_column_letter(new_idx + col_offset - 1)].width = 15
        ws.column_dimensions[get_column_letter(1)].width = 30

    right_alloc_col = len(equity_cols_to_keep) + col_offset
    ws.cell(row=row, column=right_alloc_col, value="% Alloc") 
    ws.cell(row=row, column=right_alloc_col).fill = header_fill
    ws.cell(row=row, column=right_alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=right_alloc_col).border = thin_border

    row += 1

    for _, data_row in debt_etf.iterrows():
        for old_idx in equity_cols_to_keep:
            new_idx = equity_col_map[old_idx]
            value = data_row[old_idx]
            if not pd.isna(value):
                cell_value = format_num(value)
                ws.cell(row=row, column=new_idx + col_offset - 1, value=cell_value)
                if old_idx == 0:  
                    ws.cell(row=row, column=new_idx + col_offset - 1).alignment = Alignment(horizontal="left")
                else:
                    ws.cell(row=row, column=new_idx + col_offset - 1).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=new_idx + col_offset - 1).border = thin_border
        row += 1
    
    ws.cell(row=row, column=col_offset, value="Total:")
    ws.cell(row=row, column=col_offset).font = Font(bold=True)
    ws.cell(row=row, column=col_offset).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=col_offset).border = thin_border
    ws.cell(row=row, column=col_offset).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    for col_idx in [col_offset + 1, col_offset + 2, col_offset + 3,  col_offset + 5]:
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
        ws.cell(row=row, column=col_idx).border = thin_border
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    right_market_value_col = col_offset + 5
    right_pnl_col = col_offset + 4
    right_market_col_letter = get_column_letter(right_market_value_col)
    right_pnl_col_letter = get_column_letter(right_pnl_col)

    first_data_row = row - len(debt_etf)
    last_data_row = row - 1

    if first_data_row <= last_data_row:
        market_value_cell_references = [f"{right_market_col_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        market_value_formula = "=" + "+".join(market_value_cell_references)
    else:
        market_value_formula = "=0"

    ws.cell(row=row, column=right_market_value_col, value=market_value_formula)
    ws.cell(row=row, column=right_market_value_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=right_market_value_col).border = thin_border
    ws.cell(row=row, column=right_market_value_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=right_market_value_col).number_format = '#,##,##0'
    
    debt_etf_total_row = row
    market_value_total_rows['debt_etf'] = (right_market_value_col, row)
     
    if first_data_row <= last_data_row:
        pnl_cell_references = [f"{right_pnl_col_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        pnl_formula = "=" + "+".join(pnl_cell_references)
    else:
        pnl_formula = "=0"

    ws.cell(row=row, column=right_pnl_col, value=pnl_formula)
    ws.cell(row=row, column=right_pnl_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=right_pnl_col).border = thin_border
    ws.cell(row=row, column=right_pnl_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=right_pnl_col).number_format = '#,##,##0'

    ws.cell(row=row, column=right_alloc_col, value=f"={right_market_col_letter}{row}/B4*100")
    ws.cell(row=row, column=right_alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=right_alloc_col).border = thin_border
    ws.cell(row=row, column=right_alloc_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=right_alloc_col).number_format = "0.00\%"    
        
    row += 1

    row += 1
    ws.cell(row=row, column=col_offset, value="Debt Mutual Fund").font = Font(bold=True)
    row += 1

    for old_idx in mf_cols_to_keep:
        new_idx = mf_col_map[old_idx]
        header = mf_header[old_idx]
    
        if old_idx in mf_rename:
            header = mf_rename[old_idx]
        
        ws.cell(row=row, column=new_idx + col_offset - 1, value=header)
        ws.cell(row=row, column=new_idx + col_offset - 1).fill = header_fill
        ws.cell(row=row, column=new_idx + col_offset - 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=new_idx + col_offset - 1).border = thin_border

    ws.cell(row=row, column=right_alloc_col, value="% Alloc")
    ws.cell(row=row, column=right_alloc_col).fill = header_fill
    ws.cell(row=row, column=right_alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=right_alloc_col).border = thin_border

    row += 1

    for _, data_row in debt_mf.iterrows():
        for old_idx in mf_cols_to_keep:
            new_idx = mf_col_map[old_idx]
            value = data_row[old_idx]
            if not pd.isna(value):
                cell_value = format_num(value)
                ws.cell(row=row, column=new_idx + col_offset - 1, value=cell_value)
                if old_idx == 1:  
                    ws.cell(row=row, column=new_idx + col_offset - 1).alignment = Alignment(horizontal="left")
                else:
                    ws.cell(row=row, column=new_idx + col_offset - 1).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=new_idx + col_offset - 1).border = thin_border
        row += 1
    
    ws.cell(row=row, column=col_offset, value="Total:")
    ws.cell(row=row, column=col_offset).font = Font(bold=True)
    ws.cell(row=row, column=col_offset).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=col_offset).border = thin_border
    ws.cell(row=row, column=col_offset).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    for col_idx in [col_offset + 1, col_offset + 2, col_offset + 3]:
        ws.cell(row=row, column=col_idx).border = thin_border
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    right_mf_market_col = col_offset + 5
    right_mf_pnl_col = col_offset + 4
    right_mf_market_letter = get_column_letter(right_mf_market_col)
    right_mf_pnl_letter = get_column_letter(right_mf_pnl_col)

    first_data_row = row - len(debt_mf)
    last_data_row = row - 1

    if first_data_row <= last_data_row:
        market_value_cell_references = [f"{right_mf_market_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        market_value_formula = "=" + "+".join(market_value_cell_references)
    
        pnl_cell_references = [f"{right_mf_pnl_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        pnl_formula = "=" + "+".join(pnl_cell_references)
    else:
        market_value_formula = "=0"
        pnl_formula = "=0"

    debt_mf_total_row = row
    market_value_total_rows['debt_mf'] = (right_mf_market_col, row)

    
    ws.cell(row=row, column=right_mf_market_col, value=market_value_formula)
    ws.cell(row=row, column=right_mf_market_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=right_mf_market_col).border = thin_border
    ws.cell(row=row, column=right_mf_market_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=right_mf_market_col).number_format = '#,##,##0'

    ws.cell(row=row, column=right_mf_pnl_col, value=pnl_formula)
    ws.cell(row=row, column=right_mf_pnl_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=right_mf_pnl_col).border = thin_border
    ws.cell(row=row, column=right_mf_pnl_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=right_mf_pnl_col).number_format = '#,##,##0'

    ws.cell(row=row, column=right_alloc_col, value=f"={right_mf_market_letter}{row}/B4*100")
    ws.cell(row=row, column=right_alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=right_alloc_col).border = thin_border
    ws.cell(row=row, column=right_alloc_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=right_alloc_col).number_format = "0.00\%"
    
    row += 1

    row += 2
    ws.cell(row=row, column=col_offset, value="BONDS").font = Font(bold=True)
    row += 1

    for old_idx in bond_cols_to_keep:
        new_idx = equity_col_map[old_idx]  
        header = bond_header[old_idx]
    
        if old_idx in bond_rename:
            header = bond_rename[old_idx]
        
        ws.cell(row=row, column=new_idx + col_offset - 1, value=header)
        ws.cell(row=row, column=new_idx + col_offset - 1).fill = header_fill
        ws.cell(row=row, column=new_idx + col_offset - 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=new_idx + col_offset - 1).border = thin_border

    ws.cell(row=row, column=right_alloc_col, value="% Alloc")
    ws.cell(row=row, column=right_alloc_col).fill = header_fill
    ws.cell(row=row, column=right_alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=right_alloc_col).border = thin_border

    row += 1

    for _, data_row in bond_data.iterrows():
        for old_idx in bond_cols_to_keep:
            new_idx = equity_col_map[old_idx]
            value = data_row[old_idx]
            if not pd.isna(value):
                cell_value = format_num(value)
                ws.cell(row=row, column=new_idx + col_offset - 1, value=cell_value)
                ws.cell(row=row, column=new_idx + col_offset - 1).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=new_idx + col_offset - 1).border = thin_border
        row += 1
    
    ws.cell(row=row, column=col_offset, value="Total:")
    ws.cell(row=row, column=col_offset).font = Font(bold=True)
    ws.cell(row=row, column=col_offset).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=col_offset).border = thin_border
    ws.cell(row=row, column=col_offset).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    for col_idx in [col_offset + 1, col_offset + 2, col_offset + 3]:
        ws.cell(row=row, column=col_idx).border = thin_border
        ws.cell(row=row, column=col_idx).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    bond_market_col = col_offset + 5
    bond_pnl_col = col_offset + 4
    bond_market_letter = get_column_letter(bond_market_col)
    bond_pnl_letter = get_column_letter(bond_pnl_col)

    first_data_row = row - len(bond_data)
    last_data_row = row - 1

    if first_data_row <= last_data_row:
        market_value_cell_references = [f"{bond_market_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        market_value_formula = "=" + "+".join(market_value_cell_references)
    
        pnl_cell_references = [f"{bond_pnl_letter}{r}" for r in range(first_data_row, last_data_row + 1)]
        pnl_formula = "=" + "+".join(pnl_cell_references)
    else:
        market_value_formula = "=0"
        pnl_formula = "=0"

    bond_total_row = row
    market_value_total_rows['bond'] = (bond_market_col, row)

    
    ws.cell(row=row, column=bond_market_col, value=market_value_formula)
    ws.cell(row=row, column=bond_market_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=bond_market_col).border = thin_border
    ws.cell(row=row, column=bond_market_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=bond_market_col).number_format = '#,##,##0'

    ws.cell(row=row, column=bond_pnl_col, value=pnl_formula)
    ws.cell(row=row, column=bond_pnl_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=bond_pnl_col).border = thin_border
    ws.cell(row=row, column=bond_pnl_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=bond_pnl_col).number_format = '#,##,##0'

    ws.cell(row=row, column=right_alloc_col, value=f"={bond_market_letter}{row}/B4*100")
    ws.cell(row=row, column=right_alloc_col).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=right_alloc_col).border = thin_border
    ws.cell(row=row, column=right_alloc_col).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    ws.cell(row=row, column=right_alloc_col).number_format = "0.00\%"
        
    row += 2

    portfolio_formula_parts = []
    for asset_type, (col, row_num) in market_value_total_rows.items():
        col_letter = get_column_letter(col)
        portfolio_formula_parts.append(f"{col_letter}{row_num}")

    portfolio_formula_parts.append(f"B{row_cash}")  

    portfolio_formula = "=" + "+".join(portfolio_formula_parts)
    
    ws.cell(row=4, column=2, value=portfolio_formula)
    
    wb.formula_attributes = {'calculate': 'on_load'}
    
    ws['F11'].number_format = '#,##,##0'
    
    output_filename = f"{client_code}_Portfolio.xlsx"
    wb.save(output_filename)
    
    return output_filename